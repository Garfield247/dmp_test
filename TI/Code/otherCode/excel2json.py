import json

from openpyxl import load_workbook

file = "1.xlsx"
un_use_str = '123.'


class ExcelUtils:
    def __init__(self):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    # 行数
    def get_rows(self):
        rows = self.ws.max_row
        return rows

    # 列数
    def get_clos(self):
        clo = self.ws.max_column
        return clo

    # 获取值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 修改值并保存
    def set_cell_value(self, row, column, cell_value):
        try:
            self.ws.cell(row=row, column=column).value = cell_value
            self.wb.save(self.file)
        except Exception as e:
            print("error :{}".format(e))
            self.wb.save(self.file)

    # 替换单元格中的内容
    def replace_cell_value(self):
        # 遍历第一行的值，
        for i in range(1, self.get_clos() + 1):
            cell_value = self.get_cell_value(1, i)
            # 是否存在需要替换的值
            if un_use_str in cell_value:
                cell_replace = cell_value.replace(un_use_str, "")
                self.set_cell_value(1, i, cell_replace)


def to_json():
    excel_utils = ExcelUtils()
    excel_dict = {}

    clo = excel_utils.get_clos()
    # 遍历excel中的值存入字典中
    for i in range(1, clo + 1):
        dict_key = excel_utils.get_cell_value(1, i)
        dict_value = excel_utils.get_cell_value(2, i)
        excel_dict[dict_key] = dict_value
    # 字典转json
    excel_json = json.dumps(excel_dict)
    return excel_json


if __name__ == "__main__":
    compare_json = to_json()
    print(compare_json)
